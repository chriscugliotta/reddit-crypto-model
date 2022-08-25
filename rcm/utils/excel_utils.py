import logging
import openpyxl
import pandas as pd
from copy import copy
from pandas import DataFrame
from pathlib import Path
from typing import Tuple
log = logging.getLogger(__name__)



def to_excel(
        template_file: Path,
        output_file: Path,
        sheet_name: str,
        top_left: Tuple[int, int],
        df: DataFrame,
        fill_down_styles: bool = False,
        fill_down_formulas: bool = False,
    ):
    """
    Writes a Pandas dataframe to an XLSX file.

    Args:
        template_file (Path):
            An existing (usually empty, but not necessarily) XLSX file to be populated.

        output_file (Path):
            After populating the template file, we save it as this.

        sheet_name (str):
            The name of the sheet to populate.

        top_left (tuple):
            The top-left coordinate of the range to populate, e.g. (2, 1) represents cell A2.

        df (DataFrame):
            The data to write.

        fill_down_styles (bool):
            If true, all cell styles (in first row of populated region) will be filled downward
            (throughout the populated region).

        fill_down_formulas (bool):
            If true, all formulas (adjacent to the populated region) will be filled downward
            (alongside the populated region).

    Note:
        When populating an Excel column, all cell styling (e.g. number format, font, coloring,
        etc.) will be inherited from the column's top (non-header) cell.  This way, you can
        define and store all formatting rules within the template file itself.  For example,
        when creating a template file, it is recommended to prepare a single row of dummy data,
        with all styling rules applied.  Then, when the template is filled with real data, the
        dummy values will be overwritten, but the styling rules will persist, and will propagate
        downward across the entire inbound data set.
    """

    # Read workbook and worksheet into memory.
    workbook = openpyxl.load_workbook(template_file)
    worksheet = workbook[sheet_name]

    # Loop over dataframe columns and rows.
    for j, column_name in enumerate(df.columns):
        for i in range(df.shape[0]):

            # Add dataframe value to corresponding Excel cell.
            i2 = top_left[0] + i
            j2 = top_left[1] + j
            value = df.iloc[i, j]
            value = '' if pd.isnull(value) else value
            cell = worksheet.cell(i2, j2, value)

            # Add top cell's style to cell.
            if fill_down_styles:
                if i == 0:
                    column_top_cell = worksheet.cell(i2, j2)
                else:
                    cell._style = copy(column_top_cell._style)

    if fill_down_formulas:
        _fill_down_formulas(worksheet, top_left)

    # Save to XLSX
    workbook.save(output_file)
    log.info('output_file = {0}, sheet_name = {1}, rows = {2:,}'.format(output_file, sheet_name, df.shape[0]))


def _fill_down_formulas(worksheet, top_left):

    # Check all columns in worksheet.
    for column in worksheet.columns:

        # Get top cell of recently-populated region (within this column).
        top_cell = column[top_left[0] - 1]

        # Is this top cell a formula?
        if type(top_cell.value) is str and top_cell.value.startswith('='):
            translator = openpyxl.formula.translate.Translator(top_cell.value, origin=top_cell.coordinate)

            # If so, we will propagate the top cell's formula (and style) downward.
            for cell in column[top_left[0]:]:
                cell = worksheet.cell(cell.row, cell.column, translator.translate_formula(cell.coordinate))
                cell._style = copy(top_cell._style)
